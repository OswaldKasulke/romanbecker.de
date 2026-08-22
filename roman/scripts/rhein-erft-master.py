#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Baut data/strassen-rhein-erft.json aus der Strassenuebersicht des Rhein-Erft-Kreises.

Quelle : Strassenuebersicht_Rhein-Erft-Kreis_<Stand>.xlsx, Blatt "Straßenverzeichnis"
         abgeleitet aus den amtlichen Gebaeudereferenzen NRW (Geobasis NRW,
         Datenlizenz Deutschland Zero 2.0), PLZ ueber Punkt-in-Polygon und, wo
         vorhanden, gegen die kommunalen D28-Verzeichnisse geprueft.
Aufruf : python3 scripts/rhein-erft-master.py <pfad-zur-xlsx>

Grenze des Datensatzes, die auf jeder Auswertung mitgedacht werden muss:
Gebaeudereferenzen bilden Hausnummern ab, keine Strassenregister — Strassen
ohne jede Hausnummer fehlen. Das Koelner Strassenverzeichnis
(data/strassen-master.json) ist dagegen ein vollstaendiges Register. Die beiden
Datensaetze sind deshalb nicht ohne Weiteres vergleichbar.
"""
import json, re, sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
QUELLE = Path(sys.argv[1]) if len(sys.argv) > 1 else \
    Path.home() / 'Downloads' / 'Strassenuebersicht_Rhein-Erft-Kreis_2026-01-01.xlsx'
ZIEL = ROOT / 'data' / 'strassen-rhein-erft.json'

if not QUELLE.exists():
    sys.exit(f'Datei nicht gefunden: {QUELLE}')

wb = openpyxl.load_workbook(QUELLE, read_only=True, data_only=True)
ws = wb['Straßenverzeichnis']

# Kopfzeile suchen statt auf eine feste Zeilennummer zu setzen
kopf, start = None, None
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
    if row and row[0] == 'AGS Gemeinde':
        kopf, start = [c.strip() if isinstance(c, str) else c for c in row], i + 1
        break
if not kopf:
    sys.exit('Kopfzeile "AGS Gemeinde" im Blatt Straßenverzeichnis nicht gefunden')
sp = {name: n for n, name in enumerate(kopf) if name}

def txt(w):
    return None if w is None else str(w).strip() or None

def zahl(w):
    if w is None:
        return None
    m = re.search(r'\d+', str(w))
    return int(m.group()) if m else None

def spanne(von, bis):
    v, b = zahl(von), zahl(bis)
    return [v, b] if v is not None and b is not None else None

zeilen = []
for row in ws.iter_rows(min_row=start, values_only=True):
    gemeinde = txt(row[sp['Gemeinde']])
    strasse = txt(row[sp['Straße']])
    if not gemeinde or not strasse:
        continue
    zeilen.append({
        's': strasse,
        'gem': gemeinde,
        'ot': txt(row[sp['Ortsteil']]),
        'p': txt(row[sp['PLZ']]),
        'u': spanne(row[sp['ungerade von']], row[sp['ungerade bis']]),
        'g': spanne(row[sp['gerade von']], row[sp['gerade bis']]),
        'sl': txt(row[sp['Straßenschlüssel']]),
        'pq': txt(row[sp['PLZ-Prüfstatus']]),
    })

zeilen.sort(key=lambda r: (r['gem'], r['s'].lower(), r['p'] or ''))
ZIEL.write_text(json.dumps(zeilen, ensure_ascii=False, indent=1) + '\n', encoding='utf-8')

gemeinden = sorted({r['gem'] for r in zeilen})
print(f"{ZIEL.relative_to(ROOT)} geschrieben")
print(f"  Quelle        : {QUELLE.name}")
print(f"  Zeilen        : {len(zeilen)}")
print(f"  Gemeinden     : {len(gemeinden)} ({', '.join(gemeinden)})")
print(f"  Straßennamen  : {len({r['s'] for r in zeilen})}")
print(f"  Gemeinde+Str. : {len({(r['gem'], r['s']) for r in zeilen})}")
print(f"  ohne Ortsteil : {sum(1 for r in zeilen if not r['ot'])}")
print(f"  ohne Nummern  : {sum(1 for r in zeilen if not r['u'] and not r['g'])}")
